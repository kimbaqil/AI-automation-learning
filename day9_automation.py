from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb = load_workbook("supplier_report.xlsx")
ws = wb["Raw Data"]
green_fill = PatternFill("solid", fgColor="1D9E75")
white_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = green_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center")
wb.save("supplier_report_formatted.xlsx")
print("Header formatted.")

for col in ws.columns:
    max_len = 0
col_letter = col[0].column_letter
for cell in col:
    if cell.value:
        max_len = max(max_len, len(str(cell.value)))
ws.column_dimensions[col_letter].width = max_len + 4
wb.save("supplier_report_formatted.xlsx")
print("Columns resized.")

red_fill = PatternFill("solid", fgColor="FFCCCC")
for row in ws.iter_rows(min_row=2):
    risk_cell = row[3]
    if risk_cell.value == "High":
        for cell in row:
            cell.fill = red_fill
wb.save("supplier_report_formatted.xlsx")
print("High risk rows highlighted.")

from openpyxl.styles import Border, Side
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows():
    for cell in row:
        cell.border = border
wb.save("supplier_report_formatted.xlsx")
print("Borders added.")

ws.freeze_panes = "A2"
wb.save("supplier_report_formatted.xlsx")
print("Header row frozen.")
