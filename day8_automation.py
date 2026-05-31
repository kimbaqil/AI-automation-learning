import pandas as pd
import matplotlib.pyplot as plt

df = pd.read_excel("suppliers.xlsx")

plt.figure(figsize=(8, 4))
colors = ["#c0392b" if s == "High" else "#27ae60" for s in df["risk_level"]]
plt.bar(df["supplier"], df["score"], color=colors)
plt.title("Supplier Scores by Risk Level")
plt.xlabel("Supplier")
plt.ylabel("Score")
plt.ylim(0, 100)
plt.tight_layout()
plt.savefig("supplier_scores.png", dpi=150)
plt.show()
print("Chart saved.")

status_counts = df["status"].value_counts()
plt.figure(figsize=(5, 5))
plt.pie(status_counts, labels=status_counts.index,
autopct="%1.0f%%", startangle=90)
plt.title("Supplier Status Distribution")
plt.tight_layout()
plt.savefig("status_pie.png", dpi=150)
plt.show()

fig, axes = plt.subplots(2, 2, figsize=(12, 8))
fig.suptitle("Supplier KPI Dashboard", fontsize=14, fontweight="bold")
axes[0, 0].bar(df["supplier"], df["score"])
axes[0, 0].set_title("Score by Supplier")
axes[0, 0].set_ylim(0, 100)
counts = df["status"].value_counts()
axes[0, 1].pie(counts, labels=counts.index, autopct="%1.0f%%")
axes[0, 1].set_title("Status Distribution")
risk_counts = df["risk_level"].value_counts()
axes[1, 0].barh(risk_counts.index, risk_counts.values)
axes[1, 0].set_title("Risk Level Count")
axes[1, 1].scatter(df["supplier"], df["score"], s=100)
axes[1, 1].set_title("Score Scatter")
axes[1, 1].set_ylim(0, 100)
plt.tight_layout()
plt.savefig("kpi_dashboard.png", dpi=150)
plt.show()
print("Dashboard saved.")

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
wb = load_workbook("supplier_report.xlsx")
ws = wb.create_sheet("Dashboard")
img = Image("kpi_dashboard.png")
img.anchor = "B2"
ws.add_image(img)
wb.save("supplier_report_v2.xlsx")
print("Dashboard added to Excel.")

import seaborn as sns
sns.set_theme(style="whitegrid")
plt.figure(figsize=(8, 4))
sns.barplot(data=df, x="supplier", y="score", hue="risk_level",
palette={"Low": "#27ae60", "Medium": "#f39c12", "High": "#c0392b"})
plt.title("Supplier Scores — seaborn style")
plt.tight_layout()
plt.savefig("supplier_seaborn.png", dpi=150)
plt.show()